import sys
sys.stdout.reconfigure(encoding='utf-8')
try:
    import openpyxl
    wb = openpyxl.load_workbook('data/用户/管理员教师名单.xlsx')
    print('Teacher roster OK')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f'  Sheet: {sheet}, rows={ws.max_row}, cols={ws.max_column}')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            print(f'    {row}')
except Exception as e:
    print(f'Error: {e}')

try:
    wb2 = openpyxl.load_workbook('data/CKGDT/用户/学生名单.xlsx')
    print('\nStudent roster OK')
    for sheet in wb2.sheetnames:
        ws = wb2[sheet]
        print(f'  Sheet: {sheet}, rows={ws.max_row}, cols={ws.max_column}')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            print(f'    {row}')
except Exception as e:
    print(f'Student roster error: {e}')
